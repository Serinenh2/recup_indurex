"""
Importe la nomenclature complète des déchets depuis un fichier Excel.
Fichier source : nomenclature_dechets.xlsx (Décret exécutif n°06-104)

Usage :
    python manage.py import_nomenclature_excel /chemin/vers/fichier.xlsx
"""
import os
import openpyxl
from django.core.management.base import BaseCommand, CommandError
from apps.nomenclature.models import Nomenclature


CLASSE_MAP = {
    'MA': 'MA',
    'I': 'I',
    'S': 'S',
    'SD': 'SD',
    'Ménagers et Assimilés': 'MA',
    'Inertes': 'I',
    'Spéciaux': 'S',
    'Spéciaux Dangereux': 'SD',
}

DANGEROUS_CLASSES = {'S', 'SD'}

# Map French danger keywords to model boolean fields
DANGER_KEYWORDS = {
    'toxique': 'toxique',
    'cancérogène': 'cancerogene',
    'cancerogene': 'cancerogene',
    'inflammable': 'inflammable',
    'explosible': 'explosible',
    'corrosif': 'corrosive',
    'corrosive': 'corrosive',
    'infectieux': 'infectieuse',
    'infectieuse': 'infectieuse',
    'dangereuse pour l\'environnement': 'dangereuse_environnement',
    'dangereuse pour lenvironnement': 'dangereuse_environnement',
}


def normalize_code(raw_code):
    """Normalize code like '1.3.1' → '01.03.01'"""
    parts = raw_code.strip().split('.')
    if len(parts) == 3:
        return f"{int(parts[0]):02d}.{int(parts[1]):02d}.{int(parts[2]):02d}"
    return raw_code.strip()


def extract_famille(code):
    """Extract famille from normalized code: '01.03.01' → '01'"""
    return code.split('.')[0]


def extract_sous_famille(code):
    """Extract sous_famille from normalized code: '01.03.01' → '01 03'"""
    parts = code.split('.')
    return f"{parts[0]} {parts[1]}"


def parse_danger_flags(dangerosite_fr):
    """Parse danger criteria string into boolean flags."""
    if not dangerosite_fr:
        return {}
    text = dangerosite_fr.lower().strip()
    flags = {}
    for keyword, field in DANGER_KEYWORDS.items():
        if keyword in text:
            flags[field] = True
    return flags


class Command(BaseCommand):
    help = 'Importe la nomenclature des déchets depuis un fichier Excel (.xlsx)'

    def add_arguments(self, parser):
        parser.add_argument('fichier', type=str, help='Chemin vers le fichier Excel (.xlsx)')

    def handle(self, *args, **options):
        filepath = options['fichier']

        if not os.path.exists(filepath):
            raise CommandError(f"Fichier introuvable : {filepath}")

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        except Exception as e:
            raise CommandError(f"Erreur lecture Excel : {e}")

        ws = wb.active

        # Find header row (contains "Code du déchet" or "رمز النفاية")
        header_row = None
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            for cell in row:
                if cell and ('Code du déchet' in str(cell) or 'رمز النفاية' in str(cell)):
                    header_row = i
                    break
            if header_row:
                break

        if not header_row:
            raise CommandError("En-tête non trouvée dans le fichier Excel")

        created = 0
        updated = 0
        errors = 0
        skipped = 0

        for row_num, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            if not row or not row[0]:
                skipped += 1
                continue

            raw_code = str(row[0]).strip()
            if not raw_code or not any(c.isdigit() for c in raw_code):
                skipped += 1
                continue

            try:
                code = normalize_code(raw_code)
                designation_ar = str(row[1] or '').strip()
                designation_fr = str(row[2] or '').strip()
                classe_raw = str(row[4] or '').strip()
                dangerosite_fr = str(row[6] or '').strip()
                dangerosite_ar = str(row[5] or '').strip()
                annexe = str(row[7] or '').strip()

                # Normalize classe
                classe = CLASSE_MAP.get(classe_raw)
                if not classe:
                    # Try to match partial
                    for key, val in CLASSE_MAP.items():
                        if key.lower() in classe_raw.lower():
                            classe = val
                            break
                if not classe:
                    self.stdout.write(self.style.WARNING(
                        f"Ligne {row_num}: Classe inconnue '{classe_raw}' pour code {code}, ignoré"
                    ))
                    errors += 1
                    continue

                # Determine BSD and agrément based on classe and annexe
                bsd_obligatoire = classe in DANGEROUS_CLASSES
                agrement_requis = classe in DANGEROUS_CLASSES

                # Parse danger flags
                danger_flags = parse_danger_flags(dangerosite_fr)

                # Build defaults
                defaults = {
                    'famille': extract_famille(code),
                    'sous_famille': extract_sous_famille(code),
                    'designation_fr': designation_fr,
                    'designation_ar': designation_ar,
                    'classe': classe,
                    'dangerosite_fr': dangerosite_fr,
                    'dangerosite_ar': dangerosite_ar,
                    'annexe': annexe,
                    'bsd_obligatoire': bsd_obligatoire,
                    'agrement_requis': agrement_requis,
                    **danger_flags,
                }

                obj, was_created = Nomenclature.objects.update_or_create(
                    code=code,
                    defaults=defaults
                )

                if was_created:
                    created += 1
                else:
                    updated += 1

            except Exception as e:
                self.stdout.write(self.style.ERROR(
                    f"Ligne {row_num}: Erreur pour code '{raw_code}': {e}"
                ))
                errors += 1

        wb.close()

        self.stdout.write(self.style.SUCCESS(
            f"\nImport terminé !\n"
            f"  Créés   : {created}\n"
            f"  Mis à jour : {updated}\n"
            f"  Erreurs : {errors}\n"
            f"  Ignorés : {skipped}\n"
            f"  Total   : {created + updated}"
        ))
