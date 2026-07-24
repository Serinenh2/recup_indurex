"""
Importe le glossaire des déchets depuis un fichier Excel.
Ajoute les entrées sans écraser le glossaire existant.

Usage :
    python manage.py import_glossaire_excel /chemin/vers/fichier.xlsx
"""
import os
import openpyxl
from django.core.management.base import BaseCommand, CommandError
from apps.ai_assistant.glossaire_data import GLOSSAIRE


# Glossary entries are in a Python list in glossaire_data.py.
# This command reads the Excel and appends new entries to that list at runtime,
# then writes them back to the file.

GLOSSAIRE_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    '..', '..', 'glossaire_data.py'
)


def parse_excel_glossaire(filepath):
    """Parse glossaire Excel file and return list of dicts."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    # Find header row
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        for cell in row:
            if cell and ('Terme' in str(cell) or 'المصطلح' in str(cell)):
                header_row = i
                break
        if header_row:
            break

    if not header_row:
        raise CommandError("En-tête non trouvée (colonne 'Terme' manquante)")

    entries = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or not row[0]:
            continue
        terme_fr = str(row[0]).strip()
        terme_ar = str(row[1] or '').strip()
        definition_fr = str(row[2] or '').strip()
        definition_ar = str(row[3] or '').strip()
        reference = str(row[4] or '').strip()

        if terme_fr and definition_fr:
            entries.append({
                'terme_fr': terme_fr,
                'terme_ar': terme_ar,
                'definition_fr': definition_fr,
                'definition_ar': definition_ar,
                'reference': reference,
                'categorie': 'emballage',
                'classe': 'MA',
            })

    wb.close()
    return entries


class Command(BaseCommand):
    help = 'Importe le glossaire des déchets depuis un fichier Excel (.xlsx)'

    def add_arguments(self, parser):
        parser.add_argument('fichier', type=str, help='Chemin vers le fichier Excel (.xlsx)')

    def handle(self, *args, **options):
        filepath = options['fichier']

        if not os.path.exists(filepath):
            raise CommandError(f"Fichier introuvable : {filepath}")

        new_entries = parse_excel_glossaire(filepath)

        if not new_entries:
            self.stdout.write(self.style.WARNING("Aucune entrée trouvée dans le fichier"))
            return

        # Check which entries already exist (by terme_fr)
        existing_terms = {e['terme_fr'].lower() for e in GLOSSAIRE}
        added = 0
        skipped = 0

        for entry in new_entries:
            if entry['terme_fr'].lower() in existing_terms:
                skipped += 1
                continue
            GLOSSAIRE.append(entry)
            existing_terms.add(entry['terme_fr'].lower())
            added += 1

        # Write back to glossaire_data.py
        glossaire_path = os.path.normpath(GLOSSAIRE_FILE)
        if not os.path.exists(glossaire_path):
            raise CommandError(f"Fichier glossaire introuvable : {glossaire_path}")

        with open(glossaire_path, 'r', encoding='utf-8') as f:
            content = f.read()

        # Find the GLOSSAIRE list and append new entries before the closing bracket
        import json
        new_entries_str = json.dumps(new_entries, ensure_ascii=False, indent=4)

        # Find the last ']' in the file (end of GLOSSAIRE list)
        last_bracket = content.rfind(']')
        if last_bracket == -1:
            raise CommandError("Impossible de trouver la fin de la liste GLOSSAIRE")

        # Insert new entries before the last ']'
        insert_text = ',\n'.join([
            json.dumps(e, ensure_ascii=False) for e in new_entries
        ])
        new_content = content[:last_bracket] + insert_text + ',\n' + content[last_bracket:]

        with open(glossaire_path, 'w', encoding='utf-8') as f:
            f.write(new_content)

        self.stdout.write(self.style.SUCCESS(
            f"\nImport glossaire terminé !\n"
            f"  Ajoutés   : {added}\n"
            f"  Ignorés (doublons) : {skipped}\n"
            f"  Total glossaire : {len(GLOSSAIRE)}"
        ))
